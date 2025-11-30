# from typing import List, Optional 
# from fastapi import APIRouter, Depends, HTTPException, status, Query ,BackgroundTasks 
# from sqlalchemy.orm import Session ,joinedload
# from fastapi.responses import StreamingResponse
# from openpyxl import Workbook
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from openpyxl.utils import get_column_letter
# import io
# from datetime import datetime
# import re
# import logging
# from app.database.connection import get_db
# from app.middleware.auth_middleware import get_current_active_user, require_teacher, require_student
# from app.services.classroom_service import ClassroomService
# from app.schemas.classroom_schemas import *
# from app.models.user_models import User , Notification ,UserRole

# from app.database.quiz import Quiz
# from datetime import datetime
# from app.models.classroom_models import (
#     Classroom, ClassroomMembership, ClassroomQuizAssignment, 
#     ClassroomQuizSubmission, ClassroomStatus, MembershipStatus, AssignmentStatus
# )
# from app.schemas.classroom_schemas import AssignmentResponse, ClassroomAssignmentCreate
# from app.services.email_service import EmailService
# from app.middleware.auth_middleware import get_current_user
# import traceback



# # Configure logging
# logging.basicConfig(
#     level=logging.INFO,
#     format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
# )
# logger = logging.getLogger(__name__)



# router = APIRouter()

# @router.post("/", response_model=ClassroomResponse, status_code=status.HTTP_201_CREATED)
# async def create_classroom(
#     classroom_data: ClassroomCreate,
#     current_user: User = Depends(require_teacher),
#     db: Session = Depends(get_db)
# ):
#     """Create new classroom (teachers only)"""
#     service = ClassroomService(db)
#     classroom = service.create_classroom(current_user.id, classroom_data)
#     return ClassroomResponse.model_validate(classroom , from_attributes=True)
# @router.delete("/{classroom_id}")
# async def archive_classroom(
#     classroom_id: int,
#     current_user: User = Depends(require_teacher),
#     db: Session = Depends(get_db)
# ):
#     """Archive classroom (soft delete) - changes status to ARCHIVED"""
    
#     # Get classroom and verify ownership
#     classroom = db.query(Classroom).filter(
#         Classroom.id == classroom_id,
#         Classroom.teacher_id == current_user.id
#     ).first()
    
#     if not classroom:
#         raise HTTPException(
#             status_code=status.HTTP_404_NOT_FOUND,
#             detail="Classroom not found or you don't have permission to delete it"
#         )
    
#     # Check if classroom is already archived
#     if classroom.status == ClassroomStatus.ARCHIVED:
#         raise HTTPException(
#             status_code=status.HTTP_400_BAD_REQUEST,
#             detail="Classroom is already archived"
#         )
    
#     # Check for active assignments
#     active_assignments_count = db.query(ClassroomQuizAssignment).filter(
#         ClassroomQuizAssignment.classroom_id == classroom_id,
#         ClassroomQuizAssignment.status == AssignmentStatus.ACTIVE
#     ).count()
    
#     if active_assignments_count > 0:
#         raise HTTPException(
#             status_code=status.HTTP_400_BAD_REQUEST,
#             detail=f"Cannot archive classroom with {active_assignments_count} active assignment(s). Please complete or cancel them first."
#         )
    
#     # Archive the classroom
#     classroom.status = ClassroomStatus.ARCHIVED
#     classroom.updated_at = datetime.utcnow()
    
#     # Remove all active students (set membership to REMOVED)
#     active_memberships = db.query(ClassroomMembership).filter(
#         ClassroomMembership.classroom_id == classroom_id,
#         ClassroomMembership.status == MembershipStatus.ACTIVE
#     ).all()
    
#     for membership in active_memberships:
#         membership.status = MembershipStatus.REMOVED
#         membership.removed_at = datetime.utcnow()
    
#     # Update student count
#     classroom.student_count = 0
    
#     # Cancel any draft assignments
#     draft_assignments = db.query(ClassroomQuizAssignment).filter(
#         ClassroomQuizAssignment.classroom_id == classroom_id,
#         ClassroomQuizAssignment.status == AssignmentStatus.DRAFT
#     ).all()
    
#     for assignment in draft_assignments:
#         assignment.status = AssignmentStatus.CANCELLED
#         assignment.updated_at = datetime.utcnow()
    
#     try:
#         db.commit()
#     except Exception as e:
#         db.rollback()
#         raise HTTPException(
#             status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
#             detail="Failed to archive classroom. Please try again."
#         )
    
#     return {
#         "message": "Classroom archived successfully",
#         "classroom_id": classroom_id,
#         "classroom_name": classroom.name,
#         "students_removed": len(active_memberships),
#         "assignments_cancelled": len(draft_assignments)
#     }

# # Optional: Restore archived classroom
# @router.put("/{classroom_id}/restore")
# async def restore_classroom(
#     classroom_id: int,
#     current_user: User = Depends(require_teacher),
#     db: Session = Depends(get_db)
# ):
#     """Restore an archived classroom back to active status"""
    
#     # Get classroom and verify ownership
#     classroom = db.query(Classroom).filter(
#         Classroom.id == classroom_id,
#         Classroom.teacher_id == current_user.id
#     ).first()
    
#     if not classroom:
#         raise HTTPException(
#             status_code=status.HTTP_404_NOT_FOUND,
#             detail="Classroom not found or you don't have permission"
#         )
    
#     if classroom.status != ClassroomStatus.ARCHIVED:
#         raise HTTPException(
#             status_code=status.HTTP_400_BAD_REQUEST,
#             detail="Only archived classrooms can be restored"
#         )
    
#     # Restore classroom
#     classroom.status = ClassroomStatus.ACTIVE
#     classroom.updated_at = datetime.utcnow()
    
#     try:
#         db.commit()
#     except Exception as e:
#         db.rollback()
#         raise HTTPException(
#             status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
#             detail="Failed to restore classroom. Please try again."
#         )
    
#     return {
#         "message": "Classroom restored successfully",
#         "classroom_id": classroom_id,
#         "classroom_name": classroom.name
#     }

# # Optional: Permanent delete (use with extreme caution)
# @router.delete("/{classroom_id}/permanent")
# async def permanently_delete_classroom(
#     classroom_id: int,
#     current_user: User = Depends(require_teacher),
#     db: Session = Depends(get_db)
# ):
#     """Permanently delete classroom and ALL related data - IRREVERSIBLE"""
    
#     # Get classroom and verify ownership
#     classroom = db.query(Classroom).filter(
#         Classroom.id == classroom_id,
#         Classroom.teacher_id == current_user.id
#     ).first()
    
#     if not classroom:
#         raise HTTPException(
#             status_code=status.HTTP_404_NOT_FOUND,
#             detail="Classroom not found or you don't have permission"
#         )
    
#     # Safety check: only allow deletion if classroom is archived
#     if classroom.status != ClassroomStatus.ARCHIVED:
#         raise HTTPException(
#             status_code=status.HTTP_400_BAD_REQUEST,
#             detail="Classroom must be archived before permanent deletion. Archive it first."
#         )
    
#     # Check for any student submissions (prevent data loss)
#     submissions_count = db.query(ClassroomQuizSubmission).join(
#         ClassroomQuizAssignment,
#         ClassroomQuizSubmission.assignment_id == ClassroomQuizAssignment.id
#     ).filter(
#         ClassroomQuizAssignment.classroom_id == classroom_id
#     ).count()
    
#     if submissions_count > 0:
#         raise HTTPException(
#             status_code=status.HTTP_400_BAD_REQUEST,
#             detail=f"Cannot permanently delete classroom with {submissions_count} student submissions. This would result in data loss."
#         )
    
#     try:
#         # Due to cascade="all, delete-orphan" in the model relationships,
#         # deleting the classroom should automatically delete related records
#         db.delete(classroom)
#         db.commit()
#     except Exception as e:
#         db.rollback()
#         raise HTTPException(
#             status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
#             detail="Failed to delete classroom. Please contact support."
#         )
    
#     return {
#         "message": "Classroom permanently deleted",
#         "classroom_id": classroom_id
#     }

# # Get archived classrooms (optional - for teachers to see what they've archived)
# @router.get("/archived")
# async def get_archived_classrooms(
#     skip: int = Query(0, ge=0),
#     limit: int = Query(100, ge=1, le=1000),
#     current_user: User = Depends(require_teacher),
#     db: Session = Depends(get_db)
# ):
#     """Get teacher's archived classrooms"""
    
#     archived_classrooms = db.query(Classroom).filter(
#         Classroom.teacher_id == current_user.id,
#         Classroom.status == ClassroomStatus.ARCHIVED
#     ).order_by(Classroom.updated_at.desc()).offset(skip).limit(limit).all()
    
#     from app.schemas.classroom_schemas import ClassroomResponse
#     return [
#         ClassroomResponse.model_validate(classroom, from_attributes=True) 
#         for classroom in archived_classrooms
#     ]



# @router.get("/my-classrooms", response_model=List[ClassroomResponse])
# async def get_my_classrooms(
#     status_filter: Optional[str] = Query(None, alias="status"),
#     skip: int = Query(0, ge=0),
#     limit: int = Query(100, ge=1, le=1000),
#     current_user: User = Depends(require_teacher),
#     db: Session = Depends(get_db)
# ):
#     """Get teacher's classrooms"""
#     service = ClassroomService(db)
#     classrooms = service.get_teacher_classrooms(
#         current_user.id, 
#         status_filter, 
#         skip, 
#         limit
#     )
#     return [ClassroomResponse.model_validate(c , from_attributes=True) for c in classrooms]

# @router.post("/join", response_model=dict)
# async def join_classroom(
#     join_data: ClassroomJoinRequest,
#     current_user: User = Depends(require_student),
#     db: Session = Depends(get_db)
# ):
#     """Join classroom using join code (students only)"""
#     service = ClassroomService(db)
#     membership = service.join_classroom(current_user.id, join_data.join_code)
    
#     return {
#         "message": "Successfully joined classroom",
#         "classroom_name": membership.classroom.name,
#         "teacher_name": membership.classroom.teacher.full_name
#     }

# @router.get("/enrolled", response_model=List[StudentClassroomResponse])
# async def get_enrolled_classrooms(
#     current_user: User = Depends(require_student),
#     db: Session = Depends(get_db)
# ):
#     """Get student's enrolled classrooms"""
#     service = ClassroomService(db)
#     classrooms_data = service.get_student_classrooms(current_user.id)
    
#     return [
#        StudentClassroomResponse(
#         classroom=data["classroom"],
#         joined_at=data["joined_at"],
#         teacher_name=data["teacher_name"],
#         membership_status=data["membership_status"],  # Use from data
#         completed_assignments=data["completed_assignments"],  # Use from data
#         total_assignments=data["total_assignments"],  # Use from data
#         average_score=data["average_score"],  # Use from data
#         last_activity=data["last_activity"],  # Use from data
#         pending_assignments=data["pending_assignments"]  # Add this missing field
#     )
#     for data in classrooms_data
#     ]

# @router.get("/{classroom_id}", response_model=ClassroomResponse)
# async def get_classroom(
#     classroom_id: int,
#     current_user: User = Depends(get_current_active_user),
#     db: Session = Depends(get_db)
# ):
#     """Get classroom details"""
#     service = ClassroomService(db)
#     classroom = service.get_classroom_by_id(classroom_id, current_user.id)
#     return ClassroomResponse.model_validate(classroom ,from_attributes=True)

# @router.get("/{classroom_id}/analytics")
# async def get_classroom_analytics(
#     classroom_id: int,
#     current_user: User = Depends(require_teacher),
#     db: Session = Depends(get_db)
# ):
#     """Get classroom analytics and performance data"""
#     service = ClassroomService(db)
#     return service.get_classroom_analytics(current_user.id, classroom_id)
# # Quick production fix - in your router
# @router.get("/{classroom_id}/students")
# async def get_classroom_students(
#     classroom_id: int,
#     current_user: User = Depends(require_teacher),
#     db: Session = Depends(get_db)
# ):
#     try:
#         service = ClassroomService(db)
#         students_data = service.get_classroom_students(current_user.id, classroom_id)
#         return {"students": students_data, "total": len(students_data)}
#     except Exception as e:
#         logger.error(f"Error fetching students: {e}")
#         raise HTTPException(status_code=500, detail="Failed to fetch students")

# @router.post("/{classroom_id}/assignments", response_model=AssignmentResponse)
# async def assign_quiz(
#     classroom_id: int,
#     assignment_data: ClassroomAssignmentCreate,
#     current_user: User = Depends(require_teacher),
#     db: Session = Depends(get_db)
# ):
#     """Assign quiz to classroom"""
#     assignment_data.classroom_id = classroom_id
#     service = ClassroomService(db)
#     assignment = service.assign_quiz_to_classroom(current_user.id, assignment_data)
#     return AssignmentResponse.model_validate(assignment , from_attributes=True)

# # submisson data 
# @router.get("/{classroom_id}/assignments")
# async def get_classroom_assignments(
#     classroom_id: int,
#     current_user: User = Depends(get_current_active_user),
#     db: Session = Depends(get_db)
# ):
#     """Get all assignments for classroom - with student submission data if student"""
#     service = ClassroomService(db)
    
#     # Check if user is a student in this classroom
#     if current_user.role == UserRole.STUDENT:
#         # Use the student-specific method that includes submission data
#         assignments = service.get_classroom_assignments_for_student(classroom_id, current_user.id)
#         return assignments
#     else:
#         # For teachers, use the existing method
#         assignments = service.get_classroom_assignments(classroom_id, current_user.id)
#         return [AssignmentResponse.model_validate(a, from_attributes=True) for a in assignments]

# @router.put("/{classroom_id}", response_model=ClassroomResponse)
# async def update_classroom(
#     classroom_id: int,
#     classroom_data: ClassroomUpdate,
#     current_user: User = Depends(require_teacher),
#     db: Session = Depends(get_db)
# ):
#     """Update classroom settings"""
#     service = ClassroomService(db)
#     classroom = service.update_classroom(current_user.id, classroom_id, classroom_data)
#     return ClassroomResponse.model_validate(classroom , from_attributes=True)

# @router.delete("/{classroom_id}/students/{student_id}")
# async def remove_student(
#     classroom_id: int,
#     student_id: int,
#     current_user: User = Depends(require_teacher),
#     db: Session = Depends(get_db)
# ):
#     """Remove student from classroom"""
#     service = ClassroomService(db)
#     service.remove_student(current_user.id, classroom_id, student_id)
#     return {"message": "Student removed successfully"}

# @router.post("/{classroom_id}/leave")
# async def leave_classroom(
#     classroom_id: int,
#     current_user: User = Depends(require_student),
#     db: Session = Depends(get_db)
# ):
#     """Student leaves classroom"""
#     service = ClassroomService(db)
#     service.leave_classroom(current_user.id, classroom_id)
#     return {"message": "Left classroom successfully"}

# # Add to classroom_routes.py

# @router.get("/assignments/{assignment_id}/student-view")
# async def get_assignment_for_student(
#     assignment_id: int,
#     current_user: User = Depends(require_student),
#     db: Session = Depends(get_db)
# ):
#     """Get assignment details for student to take quiz"""
#     service = ClassroomService(db)
#     assignment_data = service.get_assignment_for_student(assignment_id, current_user.id)
#     return assignment_data

# @router.get("/assignments/student")
# async def get_student_assignments(
#     current_user: User = Depends(require_student),
#     db: Session = Depends(get_db)
# ):
#     """Get all assignments for current student"""
#     service = ClassroomService(db)
#     assignments = service.get_student_assignments(current_user.id)
#     return assignments


# # Add to app/routers/classroom_routes.py (NOT assignment_submission_router.py)
# # Fix 3: Add logging to the classroom route
# @router.post("/{classroom_id}/assignments", response_model=AssignmentResponse)
# async def assign_quiz_to_classroom(
#     classroom_id: int,
#     assignment_data: ClassroomAssignmentCreate,
#     background_tasks: BackgroundTasks,
#     current_user: User = Depends(require_teacher),
#     db: Session = Depends(get_db)
# ):
#     """Assign quiz to classroom with enhanced notifications"""
    
#     try:
#         logger.info(f"Teacher {current_user.id} assigning quiz to classroom {classroom_id}")
        
#         service = ClassroomService(db)
#         assignment = service.assign_quiz_to_classroom(current_user.id, assignment_data)
        
#         logger.info(f"Assignment created with ID: {assignment.id}")
        
#         # Queue background task for notifications
#         background_tasks.add_task(
#             send_assignment_notifications,
#             db, classroom_id, assignment.id, current_user.full_name or current_user.username
#         )
        
#         logger.info(f"Background task queued for assignment {assignment.id}")
        
#         return AssignmentResponse.model_validate(assignment, from_attributes=True)
        
#     except Exception as e:
#         logger.error(f"Error in assign_quiz_to_classroom: {str(e)}")
#         logger.error(f"Assignment creation traceback: {traceback.format_exc()}")
#         raise HTTPException(status_code=500, detail=f"Failed to create assignment: {str(e)}")

# async def send_assignment_notifications(db: Session, classroom_id: int, assignment_id: int, teacher_name: str):
#     """Enhanced assignment notifications with proper error handling"""
    
#     try:
#         logger.info(f"Starting assignment notifications for assignment {assignment_id}")
        
#         # Fix: Use joinedload to eager load related objects
#         assignment = db.query(ClassroomQuizAssignment)\
#             .options(
#                 joinedload(ClassroomQuizAssignment.classroom),
#                 joinedload(ClassroomQuizAssignment.quiz)
#             )\
#             .filter(ClassroomQuizAssignment.id == assignment_id)\
#             .first()
        
#         if not assignment:
#             logger.error(f"Assignment {assignment_id} not found")
#             return
            
#         logger.info(f"Found assignment: {assignment.title}")
        
#         # Get all active students with explicit join
#         students = db.query(User).join(
#             ClassroomMembership, 
#             User.id == ClassroomMembership.user_id
#         ).filter(
#             ClassroomMembership.classroom_id == classroom_id,
#             ClassroomMembership.status == MembershipStatus.ACTIVE,
#             ClassroomMembership.role == 'student'  # Make sure we only get students
#         ).all()
        
#         logger.info(f"Found {len(students)} students to notify")
        
#         if not students:
#             logger.warning(f"No students found in classroom {classroom_id}")
#             return
        
#         email_service = EmailService()
#         email_success_count = 0
#         notification_success_count = 0
        
#         for student in students:
#             try:
#                 logger.info(f"Processing notifications for student: {student.email}")
                
#                 # Create in-app notification first
#                 notification = Notification(
#                     user_id=student.id,
#                     type='quiz_assigned',
#                     title=f'New Quiz: {assignment.title}',
#                     message=f'{teacher_name} assigned a new quiz in {assignment.classroom.name}',
#                     data={
#                         'assignment_id': assignment_id,
#                         'classroom_id': classroom_id,
#                         'due_date': assignment.due_date.isoformat() if assignment.due_date else None
#                     }
#                 )
#                 db.add(notification)
#                 notification_success_count += 1
#                 logger.info(f"Created in-app notification for {student.email}")
                
#                 # Send email notification
#                 try:
#                     email_sent = email_service.send_assignment_email(
#                         to_email=student.email,
#                         student_name=student.full_name or student.username,
#                         assignment=assignment,
#                         teacher_name=teacher_name
#                     )
                    
#                     if email_sent:
#                         email_success_count += 1
#                         logger.info(f"Email sent successfully to {student.email}")
#                     else:
#                         logger.error(f"Failed to send email to {student.email}")
                        
#                 except Exception as email_error:
#                     logger.error(f"Email error for {student.email}: {str(email_error)}")
#                     logger.error(f"Email error traceback: {traceback.format_exc()}")
                
#             except Exception as student_error:
#                 logger.error(f"Error processing student {student.email}: {str(student_error)}")
#                 logger.error(f"Student processing traceback: {traceback.format_exc()}")
        
#         # Commit all notifications
#         try:
#             db.commit()
#             logger.info(f"Database committed successfully")
#         except Exception as commit_error:
#             logger.error(f"Database commit failed: {str(commit_error)}")
#             db.rollback()
#             raise
        
#         logger.info(f"Notification summary: {notification_success_count} in-app notifications, {email_success_count} emails sent out of {len(students)} students")
        
#     except Exception as e:
#         logger.error(f"Critical error in send_assignment_notifications: {str(e)}")
#         logger.error(f"Full traceback: {traceback.format_exc()}")
#         try:
#             db.rollback()
#         except:
#             pass
#         raise
    
    
    
# # specific route for getting classroom details for students
# @router.get("/{classroom_id}/student-view")
# async def get_classroom_student_view(
#     classroom_id: int,
#     current_user: User = Depends(require_student),
#     db: Session = Depends(get_db)
# ):
#     """Get classroom details optimized for student view"""
    
#     # Verify student access
#     membership = db.query(ClassroomMembership).filter(
#         ClassroomMembership.classroom_id == classroom_id,
#         ClassroomMembership.user_id == current_user.id,
#         ClassroomMembership.status == MembershipStatus.ACTIVE
#     ).first()
    
#     if not membership:
#         raise HTTPException(
#             status_code=status.HTTP_403_FORBIDDEN,
#             detail="You don't have access to this classroom"
#         )
    
#     # Get classroom details
#     classroom = db.query(Classroom).filter(
#         Classroom.id == classroom_id,
#         Classroom.status == ClassroomStatus.ACTIVE
#     ).first()
    
#     if not classroom:
#         raise HTTPException(
#             status_code=status.HTTP_404_NOT_FOUND,
#             detail="Classroom not found"
#         )
    
#     # Get assignments count by status for this student
#     all_assignments = db.query(ClassroomQuizAssignment).filter(
#         ClassroomQuizAssignment.classroom_id == classroom_id,
#         ClassroomQuizAssignment.status == AssignmentStatus.ACTIVE
#     ).all()
    
#     pending_count = 0
#     completed_count = 0
#     overdue_count = 0
    
#     now = datetime.utcnow()
    
#     for assignment in all_assignments:
#         submission = db.query(ClassroomQuizSubmission).filter(
#             ClassroomQuizSubmission.assignment_id == assignment.id,
#             ClassroomQuizSubmission.student_id == current_user.id
#         ).first()
        
#         if submission:
#             completed_count += 1
#         elif assignment.due_date and now > assignment.due_date:
#             overdue_count += 1
#         else:
#             pending_count += 1
    
#     return {
#         "classroom": {
#             "id": classroom.id,
#             "name": classroom.name,
#             "description": classroom.description,
#             "subject": classroom.subject,
#             "status": classroom.status.value,
#             "student_count": classroom.student_count,
#             "created_at": classroom.created_at,
#             "teacher_name": classroom.teacher.full_name or classroom.teacher.username,
#             "teacher_id": classroom.teacher_id
#         },
#         "membership": {
#             "joined_at": membership.joined_at,
#             "role": membership.role,
#             "status": membership.status.value
#         },
#         "assignment_summary": {
#             "pending": pending_count,
#             "completed": completed_count,
#             "overdue": overdue_count,
#             "total": len(all_assignments)
#         }
#     }
    
    
    
# # adding for results and deletion and updation 

# @router.get("/{classroom_id}/assignments/{assignment_id}/results", response_model=Dict[str, Any])
# async def get_assignment_results_teacher(
#     classroom_id: int,
#     assignment_id: int,
#     current_user: User = Depends(get_current_user),
#     db: Session = Depends(get_db)
# ):
#     """Get assignment results for teacher view"""
#     # Verify teacher owns the classroom
#     classroom = db.query(Classroom).filter(
#         Classroom.id == classroom_id,
#         Classroom.teacher_id == current_user.id
#     ).first()
    
#     if not classroom:
#         raise HTTPException(
#             status_code=status.HTTP_404_NOT_FOUND,
#             detail="Classroom not found"
#         )
    
#     # Get assignment
#     assignment = db.query(ClassroomQuizAssignment).filter(
#         ClassroomQuizAssignment.id == assignment_id,
#         ClassroomQuizAssignment.classroom_id == classroom_id
#     ).first()
    
#     if not assignment:
#         raise HTTPException(
#             status_code=status.HTTP_404_NOT_FOUND,
#             detail="Assignment not found"
#         )
    
#     # Get all submissions for this assignment
#     submissions = db.query(ClassroomQuizSubmission).join(
#         User, ClassroomQuizSubmission.student_id == User.id
#     ).filter(
#         ClassroomQuizSubmission.assignment_id == assignment_id
#     ).order_by(ClassroomQuizSubmission.submitted_at.desc()).all()
    
#     # Get classroom students
#     students = db.query(User).join(
#         ClassroomMembership, ClassroomMembership.student_id == User.id
#     ).filter(
#         ClassroomMembership.classroom_id == classroom_id,
#         ClassroomMembership.status == MembershipStatus.ACTIVE
#     ).all()
    
#     # Calculate statistics
#     total_students = len(students)
#     submitted_count = len(set(s.student_id for s in submissions))
#     pending_count = total_students - submitted_count

#     # Fix: Handle empty submissions list safely
#     #get_assignment_results_teacher function, fix the percentage calculation:
#         # Replace the problematic lines with this safer version:
#     if submissions and any(s.total_marks_scored is not None for s in submissions):
#         # Calculate statistics using only submissions with valid mark data
#         valid_submissions = [s for s in submissions if s.total_marks_scored is not None and s.max_possible_marks is not None]
        
#         if valid_submissions:
#             # Calculate average based on marks
#             total_marks_earned = sum(s.total_marks_scored for s in valid_submissions)
#             total_max_marks = sum(s.max_possible_marks for s in valid_submissions)
            
#             if total_max_marks > 0:
#                 average_score = (total_marks_earned / len(valid_submissions)) / (total_max_marks / len(valid_submissions)) * 100
#             else:
#                 average_score = 0
                
#             # Calculate individual percentages for min/max
#             individual_percentages = []
#             for s in valid_submissions:
#                 if s.max_possible_marks > 0:
#                     percentage = (s.total_marks_scored / s.max_possible_marks) * 100
#                     individual_percentages.append(max(0, percentage))  # Cap at 0%
            
#             if individual_percentages:
#                 highest_score = max(individual_percentages)
#                 lowest_score = min(individual_percentages)
#             else:
#                 highest_score = lowest_score = 0
#         else:
#             average_score = highest_score = lowest_score = 0
#     else:
#         # Fallback to old percentage-based calculation for legacy data
#         valid_scores = [s.score_percentage for s in submissions if s.score_percentage is not None]
#         if valid_scores:
#             average_score = sum(valid_scores) / len(valid_scores)
#             highest_score = max(valid_scores)
#             lowest_score = min(valid_scores)
#         else:
#             average_score = highest_score = lowest_score = 0
        
#     # Format student results
#     student_results = []
#     for student in students:
#         student_submission = next((s for s in submissions if s.student_id == student.id), None)
#         student_results.append({
#             "student_id": student.id,
#             "student_name": student.full_name,
#             "student_email": student.email,
#             "submitted": bool(student_submission),
#             "submitted_at": student_submission.submitted_at if student_submission else None,
#             "score_percentage": student_submission.score_percentage if student_submission else None,
#             "is_late": student_submission.is_late if student_submission else False,
#             "attempt_number": student_submission.attempt_number if student_submission else 0,
#             "time_taken_minutes": student_submission.time_taken_minutes if student_submission else None
#         })
    
#     return {
#         "assignment": {
#             "id": assignment.id,
#             "title": assignment.title,
#             "description": assignment.description,
#             "due_date": assignment.due_date,
#             "created_at": assignment.created_at,
#             "status": assignment.status.value
#         },
#         "quiz": {
#             "id": assignment.quiz.id,
#             "title": assignment.quiz.title,
#             "total_questions": assignment.quiz.total_questions,
#             "total_points": assignment.quiz.total_points
#         },
#         "statistics": {
#             "total_students": total_students,
#             "submitted_count": submitted_count,
#             "pending_count": pending_count,
#             "completion_rate": (submitted_count / total_students * 100) if total_students > 0 else 0,
#             "average_score": round(average_score, 2),
#             "highest_score": round(highest_score, 2),
#             "lowest_score": round(lowest_score, 2)
#         },
#         "student_results": student_results
#     }

# @router.delete("/{classroom_id}/assignments/{assignment_id}", response_model=Dict[str, Any])
# async def delete_assignment(
#     classroom_id: int,
#     assignment_id: int,
#     current_user: User = Depends(get_current_user),
#     db: Session = Depends(get_db)
# ):
#     """Delete an assignment (soft delete)"""
#     # Verify teacher owns the classroom
#     classroom = db.query(Classroom).filter(
#         Classroom.id == classroom_id,
#         Classroom.teacher_id == current_user.id
#     ).first()
    
#     if not classroom:
#         raise HTTPException(
#             status_code=status.HTTP_404_NOT_FOUND,
#             detail="Classroom not found"
#         )
    
#     # Get assignment
#     assignment = db.query(ClassroomQuizAssignment).filter(
#         ClassroomQuizAssignment.id == assignment_id,
#         ClassroomQuizAssignment.classroom_id == classroom_id
#     ).first()
    
#     if not assignment:
#         raise HTTPException(
#             status_code=status.HTTP_404_NOT_FOUND,
#             detail="Assignment not found"
#         )
    
#     # Check if assignment has submissions
#     submissions_count = db.query(ClassroomQuizSubmission).filter(
#         ClassroomQuizSubmission.assignment_id == assignment_id
#     ).count()
    
#     if submissions_count > 0:
#         # Soft delete - change status to INACTIVE (or whatever inactive status exists in your enum)
#         # You can also use assignment.status = AssignmentStatus.INACTIVE if that exists
#         # For now, we'll just do a soft delete by hiding it (you can modify this based on your enum)
#         try:
#             # Try common enum values for inactive/deleted status
#             if hasattr(AssignmentStatus, 'INACTIVE'):
#                 assignment.status = AssignmentStatus.INACTIVE
#             elif hasattr(AssignmentStatus, 'ARCHIVED'):
#                 assignment.status = AssignmentStatus.ARCHIVED
#             elif hasattr(AssignmentStatus, 'DISABLED'):
#                 assignment.status = AssignmentStatus.DISABLED
#             else:
#                 # If no suitable status exists, just hard delete
#                 db.delete(assignment)
#                 db.commit()
#                 return {
#                     "message": "Assignment deleted successfully",
#                     "assignment_id": assignment_id,
#                     "action": "deleted"
#                 }
            
#             assignment.updated_at = datetime.utcnow()
#             db.commit()
            
#             return {
#                 "message": "Assignment archived successfully (has submissions)",
#                 "assignment_id": assignment_id,
#                 "action": "archived"
#             }
#         except AttributeError:
#             # If status update fails, do hard delete
#             db.delete(assignment)
#             db.commit()
#             return {
#                 "message": "Assignment deleted successfully",
#                 "assignment_id": assignment_id,
#                 "action": "deleted"
#             }
#     else:
#         # Hard delete if no submissions
#         db.delete(assignment)
#         db.commit()
        
#         return {
#             "message": "Assignment deleted successfully",
#             "assignment_id": assignment_id,
#             "action": "deleted"
#         }

# @router.put("/{classroom_id}/assignments/{assignment_id}", response_model=Dict[str, Any])
# async def update_assignment(
#     classroom_id: int,
#     assignment_id: int,
#     update_data: Dict[str, Any],
#     current_user: User = Depends(get_current_user),
#     db: Session = Depends(get_db)
# ):
#     """Update assignment details"""
#     # Verify teacher owns the classroom
#     classroom = db.query(Classroom).filter(
#         Classroom.id == classroom_id,
#         Classroom.teacher_id == current_user.id
#     ).first()
    
#     if not classroom:
#         raise HTTPException(
#             status_code=status.HTTP_404_NOT_FOUND,
#             detail="Classroom not found"
#         )
    
#     # Get assignment
#     assignment = db.query(ClassroomQuizAssignment).filter(
#         ClassroomQuizAssignment.id == assignment_id,
#         ClassroomQuizAssignment.classroom_id == classroom_id
#     ).first()
    
#     if not assignment:
#         raise HTTPException(
#             status_code=status.HTTP_404_NOT_FOUND,
#             detail="Assignment not found"
#         )
    
#     # Update allowed fields
#     allowed_fields = [
#         'title', 'description', 'instructions', 'due_date', 
#         'time_limit_minutes', 'max_attempts', 'allow_late_submission',
#         'show_results_immediately', 'shuffle_questions'
#     ]
    
#     for field, value in update_data.items():
#         if field in allowed_fields and hasattr(assignment, field):
#             setattr(assignment, field, value)
    
#     assignment.updated_at = datetime.utcnow()
#     db.commit()
#     db.refresh(assignment)
    
#     return {
#         "message": "Assignment updated successfully",
#         "assignment": {
#             "id": assignment.id,
#             "title": assignment.title,
#             "description": assignment.description,
#             "due_date": assignment.due_date,
#             "updated_at": assignment.updated_at
#         }
#     }
    
    
    
    
    
# # This is to download  the result in the excel file and  other option 
# @router.get("/{classroom_id}/assignments/{assignment_id}/export-excel")
# async def export_assignment_results_excel(
#     classroom_id: int,
#     assignment_id: int,
#     export_filter: str = "all",  # "all", "submitted", "pending"
#     current_user: User = Depends(get_current_user),
#     db: Session = Depends(get_db)
# ):
#     """Export assignment results to Excel file"""
#     # Verify teacher owns the classroom
#     classroom = db.query(Classroom).filter(
#         Classroom.id == classroom_id,
#         Classroom.teacher_id == current_user.id
#     ).first()
    
#     if not classroom:
#         raise HTTPException(
#             status_code=status.HTTP_404_NOT_FOUND,
#             detail="Classroom not found"
#         )
    
#     # Get assignment
#     assignment = db.query(ClassroomQuizAssignment).filter(
#         ClassroomQuizAssignment.id == assignment_id,
#         ClassroomQuizAssignment.classroom_id == classroom_id
#     ).first()
    
#     if not assignment:
#         raise HTTPException(
#             status_code=status.HTTP_404_NOT_FOUND,
#             detail="Assignment not found"
#         )
    
#     # Get all submissions for this assignment
#     submissions = db.query(ClassroomQuizSubmission).join(
#         User, ClassroomQuizSubmission.student_id == User.id
#     ).filter(
#         ClassroomQuizSubmission.assignment_id == assignment_id
#     ).order_by(ClassroomQuizSubmission.submitted_at.desc()).all()
    
#     # Get classroom students
#     students = db.query(User).join(
#         ClassroomMembership, ClassroomMembership.student_id == User.id
#     ).filter(
#         ClassroomMembership.classroom_id == classroom_id,
#         ClassroomMembership.status == MembershipStatus.ACTIVE
#     ).order_by(User.full_name).all()
    
#     # Create submission lookup
#     submission_dict = {s.student_id: s for s in submissions}
    
#     # Filter students based on export_filter
#     filtered_students = []
#     for student in students:
#         has_submitted = student.id in submission_dict
        
#         if export_filter == "submitted" and not has_submitted:
#             continue
#         elif export_filter == "pending" and has_submitted:
#             continue
        
#         filtered_students.append(student)
    
#     # Calculate statistics
#     total_students = len(students)
#     submitted_count = len(submissions)
#     pending_count = total_students - submitted_count
    
#         # In export_assignment_results_excel function, replace the statistics calculation:
#     if submissions and any(s.total_marks_scored is not None for s in submissions):
#         valid_submissions = [s for s in submissions if s.total_marks_scored is not None and s.max_possible_marks is not None]
        
#         if valid_submissions:
#             total_marks_earned = sum(s.total_marks_scored for s in valid_submissions)
#             total_max_marks = sum(s.max_possible_marks for s in valid_submissions)
#             average_score = (total_marks_earned / len(valid_submissions)) / (total_max_marks / len(valid_submissions)) * 100 if total_max_marks > 0 else 0
            
#             individual_percentages = [(s.total_marks_scored / s.max_possible_marks) * 100 for s in valid_submissions if s.max_possible_marks > 0]
#             highest_score = max(individual_percentages) if individual_percentages else 0
#             lowest_score = min(individual_percentages) if individual_percentages else 0
#         else:
#             average_score = highest_score = lowest_score = 0
#     else:
#         # Fallback for legacy data
#         valid_scores = [s.score_percentage for s in submissions if s.score_percentage is not None]
#         if valid_scores:
#             average_score = sum(valid_scores) / len(valid_scores)
#             highest_score = max(valid_scores)
#             lowest_score = min(valid_scores)
#         else:
#             average_score = highest_score = lowest_score = 0
    
#     # Create Excel workbook
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Assignment Results"
    
#     # Define styles
#     header_font = Font(bold=True, color="FFFFFF")
#     header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
#     title_font = Font(bold=True, size=16)
#     subtitle_font = Font(bold=True, size=12)
#     center_align = Alignment(horizontal="center", vertical="center")
#     border = Border(
#         left=Side(style='thin'),
#         right=Side(style='thin'),
#         top=Side(style='thin'),
#         bottom=Side(style='thin')
#     )
    
#     # Title and metadata
#     ws.merge_cells('A1:H1')
#     ws['A1'] = f"{assignment.title} - Results Report"
#     ws['A1'].font = title_font
#     ws['A1'].alignment = center_align
    
#     ws.merge_cells('A2:H2')
#     ws['A2'] = f"Class: {classroom.name} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
#     ws['A2'].alignment = center_align
    
#     # Summary statistics
#     current_row = 4
#     ws[f'A{current_row}'] = "Summary Statistics"
#     ws[f'A{current_row}'].font = subtitle_font
#     current_row += 1
    
#     stats_data = [
#         ("Total Students", total_students),
#         ("Submitted", submitted_count),
#         ("Pending", pending_count),
#         ("Completion Rate", f"{(submitted_count/total_students*100):.1f}%" if total_students > 0 else "0%"),
#         ("Average Score", f"{average_score:.1f}%" if average_score > 0 else "N/A"),
#         ("Highest Score", f"{highest_score:.1f}%" if highest_score > 0 else "N/A"),
#         ("Lowest Score", f"{lowest_score:.1f}%" if lowest_score > 0 else "N/A")
#     ]
    
#     for label, value in stats_data:
#         ws[f'A{current_row}'] = label
#         ws[f'B{current_row}'] = value
#         ws[f'A{current_row}'].font = Font(bold=True)
#         current_row += 1
    
#     # Student results section
#     current_row += 2
#     ws[f'A{current_row}'] = "Student Results"
#     ws[f'A{current_row}'].font = subtitle_font
#     current_row += 1
    
#     # Headers
#     headers = ["Name", "Email", "Username", "Score (Marks)", "Score (%)", "Status", "Submitted At", "Late", "Time Taken (min)", "Attempts"]
#     for col, header in enumerate(headers, 1):
#         cell = ws.cell(row=current_row, column=col)
#         cell.value = header
#         cell.font = header_font
#         cell.fill = header_fill
#         cell.alignment = center_align
#         cell.border = border
    
#     current_row += 1
    
#     # Student data
#     for student in filtered_students:
#         submission = submission_dict.get(student.id)
        
#         # Format score display
#         if submission:
#             if submission.total_marks_scored is not None and submission.max_possible_marks is not None:
#                 # New format: show actual marks
#                 marks_display = f"{submission.total_marks_scored}/{submission.max_possible_marks}"
#                 percentage_display = f"{(submission.total_marks_scored / submission.max_possible_marks * 100):.1f}%" if submission.max_possible_marks > 0 else "0%"
#             else:
#                 # Fallback for legacy data
#                 marks_display = "N/A"
#                 percentage_display = f"{submission.score_percentage:.1f}%" if submission.score_percentage is not None else "Not Graded"
#         else:
#             marks_display = "Not Submitted"
#             percentage_display = "Not Submitted"
        
#         row_data = [
#             student.full_name or "N/A",
#             student.email or "N/A", 
#             student.username or "N/A",
#             marks_display,  # Actual marks column
#             percentage_display,  # Percentage column
#             "Submitted" if submission else "Pending",
#             submission.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if submission else "N/A",
#             "Yes" if submission and submission.is_late else "No",
#             str(submission.time_taken_minutes) if submission and submission.time_taken_minutes else "N/A",
#             str(submission.attempt_number) if submission else "0"
#         ]
        
#         for col, value in enumerate(row_data, 1):
#             cell = ws.cell(row=current_row, column=col)
#             cell.value = value
#             cell.border = border
#             cell.alignment = center_align
            
#             # Color coding for marks column (col 4) and percentage column (col 5)
#             if col == 4 and submission and submission.total_marks_scored is not None:
#                 # Color based on whether marks are positive/negative
#                 if submission.total_marks_scored < 0:
#                     cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red for negative
#                 elif submission.total_marks_scored >= submission.max_possible_marks * 0.8:
#                     cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green for 80%+
#                 elif submission.total_marks_scored >= submission.max_possible_marks * 0.6:
#                     cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow for 60-79%
#                 else:
#                     cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red for <60%
            
#             elif col == 5 and submission and submission.total_marks_scored is not None:
#                 # Color coding for percentage column
#                 percentage = (submission.total_marks_scored / submission.max_possible_marks * 100) if submission.max_possible_marks > 0 else 0
#                 if percentage >= 80:
#                     cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
#                 elif percentage >= 60:
#                     cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
#                 else:
#                     cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
        
#         current_row += 1
    
#     # Auto-adjust column widths
#     for column in ws.columns:
#         max_length = 0
#         column_letter = get_column_letter(column[0].column)
#         for cell in column:
#             try:
#                 if len(str(cell.value)) > max_length:
#                     max_length = len(str(cell.value))
#             except:
#                 pass
#         adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
#         ws.column_dimensions[column_letter].width = adjusted_width
    
#     # Save to memory
#     excel_buffer = io.BytesIO()
#     wb.save(excel_buffer)
#     excel_buffer.seek(0)
    
#     # Create filename (sanitize for cross-platform compatibility)
#     safe_assignment_title = re.sub(r'[^\w\s-]', '', assignment.title)
#     safe_assignment_title = re.sub(r'[-\s]+', '-', safe_assignment_title)
#     date_str = datetime.now().strftime('%Y%m%d_%H%M%S')
#     filename = f"Assignment_Results_{safe_assignment_title}_{date_str}.xlsx"
    
#     # Return streaming response
#     def iter_file():
#         excel_buffer.seek(0)
#         yield from excel_buffer
    
#     return StreamingResponse(
#         iter_file(),
#         media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#         headers={"Content-Disposition": f"attachment; filename={filename}"}
#     )
    
    
    
# # for downloading the excel file for each class their children score 

# @router.get("/{classroom_id}/export-classroom-excel")
# async def export_classroom_performance_excel(
#     classroom_id: int,
#     current_user: User = Depends(get_current_user),
#     db: Session = Depends(get_db)
# ):
#     """Export all students' performance across all assignments in Excel format"""
#     # Verify teacher owns the classroom
#     classroom = db.query(Classroom).filter(
#         Classroom.id == classroom_id,
#         Classroom.teacher_id == current_user.id
#     ).first()
    
#     if not classroom:
#         raise HTTPException(
#             status_code=status.HTTP_404_NOT_FOUND,
#             detail="Classroom not found"
#         )
    
#     # Get all students in classroom
#     students = db.query(User).join(
#         ClassroomMembership, ClassroomMembership.student_id == User.id
#     ).filter(
#         ClassroomMembership.classroom_id == classroom_id,
#         ClassroomMembership.status == MembershipStatus.ACTIVE
#     ).order_by(User.full_name).all()
    
#     # Get all assignments for this classroom
#     assignments = db.query(ClassroomQuizAssignment).filter(
#         ClassroomQuizAssignment.classroom_id == classroom_id
#     ).order_by(ClassroomQuizAssignment.created_at).all()
    
#     # Get all submissions for this classroom
#     submissions = db.query(ClassroomQuizSubmission).join(
#         ClassroomQuizAssignment,
#         ClassroomQuizSubmission.assignment_id == ClassroomQuizAssignment.id
#     ).filter(
#         ClassroomQuizAssignment.classroom_id == classroom_id
#     ).all()
    
#     # Create submission lookup: {student_id: {assignment_id: submission}}
#     submission_lookup = {}
#     for submission in submissions:
#         if submission.student_id not in submission_lookup:
#             submission_lookup[submission.student_id] = {}
#         submission_lookup[submission.student_id][submission.assignment_id] = submission
    
#     # Create Excel workbook
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Classroom Performance"
    
#     # Define styles
#     header_font = Font(bold=True, color="FFFFFF")
#     header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
#     title_font = Font(bold=True, size=16)
#     center_align = Alignment(horizontal="center", vertical="center")
#     border = Border(
#         left=Side(style='thin'), right=Side(style='thin'),
#         top=Side(style='thin'), bottom=Side(style='thin')
#     )
    
#     # Title
#     ws.merge_cells(f'A1:{get_column_letter(len(assignments) + 6)}1')
#     ws['A1'] = f"{classroom.name} - Complete Performance Report"
#     ws['A1'].font = title_font
#     ws['A1'].alignment = center_align
    
#     # Classroom info
#     ws.merge_cells(f'A2:{get_column_letter(len(assignments) + 6)}2')
#     ws['A2'] = f"Teacher: {classroom.teacher.full_name} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
#     ws['A2'].alignment = center_align
    
#     # Headers row
#     current_row = 4
#     headers = ["Student Name", "Email", "Username"]
    
#     # Add assignment titles as headers
#     for assignment in assignments:
#         headers.append(f"{assignment.title}")
    
#     headers.extend(["Overall Average", "Total Assignments", "Completed"])
    
#     # Write headers
#     for col, header in enumerate(headers, 1):
#         cell = ws.cell(row=current_row, column=col)
#         cell.value = header
#         cell.font = header_font
#         cell.fill = header_fill
#         cell.alignment = center_align
#         cell.border = border
    
#     current_row += 1
    
#     # Student data
#     for student in students:
#         student_scores = []
#         completed_count = 0
        
#         row_data = [
#             student.full_name or "N/A",
#             student.email or "N/A",
#             student.username or "N/A"
#         ]
        
#         # Add scores for each assignment
#         for assignment in assignments:
#             submission = submission_lookup.get(student.id, {}).get(assignment.id)
            
#             if submission:
#                 if submission.total_marks_scored is not None and submission.max_possible_marks is not None:
#                     # Show marks format: "8/10 (80%)"
#                     marks_display = f"{submission.total_marks_scored}/{submission.max_possible_marks}"
#                     percentage = (submission.total_marks_scored / submission.max_possible_marks * 100) if submission.max_possible_marks > 0 else 0
#                     score_display = f"{marks_display} ({percentage:.1f}%)"
#                     student_scores.append(percentage)
#                 else:
#                     # Fallback to percentage only
#                     score_display = f"{submission.score_percentage:.1f}%" if submission.score_percentage else "Graded"
#                     student_scores.append(submission.score_percentage or 0)
#                 completed_count += 1
#             else:
#                 score_display = "Not Submitted"
#                 student_scores.append(0)  # Count as 0 for average calculation
            
#             row_data.append(score_display)
        
#         # Calculate overall average
#         overall_average = sum(student_scores) / len(student_scores) if student_scores else 0
        
#         # Add summary columns
#         row_data.extend([
#             f"{overall_average:.1f}%",
#             len(assignments),
#             completed_count
#         ])
        
#         # Write student row
#         for col, value in enumerate(row_data, 1):
#             cell = ws.cell(row=current_row, column=col)
#             cell.value = value
#             cell.border = border
#             cell.alignment = center_align
            
#             # Color coding for assignment scores (columns 4 to 4+len(assignments)-1)
#             if 4 <= col <= 3 + len(assignments):
#                 if "Not Submitted" in str(value):
#                     cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
#                 elif any(char.isdigit() for char in str(value)):
#                     # Extract percentage for color coding
#                     if "(" in str(value) and "%" in str(value):
#                         try:
#                             percentage_str = str(value).split("(")[1].split("%")[0]
#                             percentage = float(percentage_str)
#                             if percentage >= 80:
#                                 cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
#                             elif percentage >= 60:
#                                 cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
#                             else:
#                                 cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
#                         except:
#                             pass
        
#         current_row += 1
    
#     # Add summary statistics
#     current_row += 2
#     ws[f'A{current_row}'] = "Classroom Statistics"
#     ws[f'A{current_row}'].font = Font(bold=True, size=14)
#     current_row += 1
    
#     # Calculate class averages for each assignment
#     for i, assignment in enumerate(assignments):
#         assignment_scores = []
#         for student in students:
#             submission = submission_lookup.get(student.id, {}).get(assignment.id)
#             if submission and submission.score_percentage:
#                 assignment_scores.append(submission.score_percentage)
        
#         avg_score = sum(assignment_scores) / len(assignment_scores) if assignment_scores else 0
#         completion_rate = len(assignment_scores) / len(students) * 100 if students else 0
        
#         ws[f'A{current_row}'] = f"{assignment.title} Average:"
#         ws[f'B{current_row}'] = f"{avg_score:.1f}% ({completion_rate:.1f}% completion)"
#         current_row += 1
    
#     # Auto-adjust column widths
#     for column in ws.columns:
#         max_length = 0
#         column_letter = get_column_letter(column[0].column)
#         for cell in column:
#             try:
#                 if len(str(cell.value)) > max_length:
#                     max_length = len(str(cell.value))
#             except:
#                 pass
#         adjusted_width = min(max_length + 2, 50)
#         ws.column_dimensions[column_letter].width = adjusted_width
    
#     # Save to memory
#     excel_buffer = io.BytesIO()
#     wb.save(excel_buffer)
#     excel_buffer.seek(0)
    
#     # Create filename
#     safe_classroom_name = re.sub(r'[^\w\s-]', '', classroom.name)
#     safe_classroom_name = re.sub(r'[-\s]+', '-', safe_classroom_name)
#     date_str = datetime.now().strftime('%Y%m%d_%H%M%S')
#     filename = f"Classroom_Performance_{safe_classroom_name}_{date_str}.xlsx"
    
#     # Return streaming response
#     def iter_file():
#         excel_buffer.seek(0)
#         yield from excel_buffer
    
#     return StreamingResponse(
#         iter_file(),
#         media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#         headers={"Content-Disposition": f"attachment; filename={filename}"}
#     )




from typing import List, Optional 
from fastapi import APIRouter, Depends, HTTPException, status, Query ,BackgroundTasks 
from sqlalchemy.orm import Session ,joinedload
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from datetime import datetime
import re
import logging
from app.database.connection import get_db
from app.middleware.auth_middleware import get_current_active_user, require_teacher, require_student
from app.services.classroom_service import ClassroomService
from app.schemas.classroom_schemas import *
from app.models.user_models import User , Notification ,UserRole

from app.database.quiz import Quiz
from datetime import datetime
from app.models.classroom_models import (
    Classroom, ClassroomMembership, ClassroomQuizAssignment, 
    ClassroomQuizSubmission, ClassroomStatus, MembershipStatus, AssignmentStatus
)
from app.schemas.classroom_schemas import AssignmentResponse, ClassroomAssignmentCreate
from app.services.email_service import EmailService
from app.middleware.auth_middleware import get_current_user
import traceback



# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)



router = APIRouter()

@router.post("/", response_model=ClassroomResponse, status_code=status.HTTP_201_CREATED)
async def create_classroom(
    classroom_data: ClassroomCreate,
    current_user: User = Depends(require_teacher),
    db: Session = Depends(get_db)
):
    """Create new classroom (teachers only)"""
    service = ClassroomService(db)
    classroom = service.create_classroom(current_user.id, classroom_data)
    return ClassroomResponse.model_validate(classroom , from_attributes=True)
@router.delete("/{classroom_id}")
async def archive_classroom(
    classroom_id: int,
    current_user: User = Depends(require_teacher),
    db: Session = Depends(get_db)
):
    """Archive classroom (soft delete) - changes status to ARCHIVED"""
    
    # Get classroom and verify ownership
    classroom = db.query(Classroom).filter(
        Classroom.id == classroom_id,
        Classroom.teacher_id == current_user.id
    ).first()
    
    if not classroom:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Classroom not found or you don't have permission to delete it"
        )
    
    # Check if classroom is already archived
    if classroom.status == ClassroomStatus.ARCHIVED:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Classroom is already archived"
        )
    
    # Check for active assignments
    active_assignments_count = db.query(ClassroomQuizAssignment).filter(
        ClassroomQuizAssignment.classroom_id == classroom_id,
        ClassroomQuizAssignment.status == AssignmentStatus.ACTIVE
    ).count()
    
    if active_assignments_count > 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot archive classroom with {active_assignments_count} active assignment(s). Please complete or cancel them first."
        )
    
    # Archive the classroom
    classroom.status = ClassroomStatus.ARCHIVED
    classroom.updated_at = datetime.utcnow()
    
    # Remove all active students (set membership to REMOVED)
    active_memberships = db.query(ClassroomMembership).filter(
        ClassroomMembership.classroom_id == classroom_id,
        ClassroomMembership.status == MembershipStatus.ACTIVE
    ).all()
    
    for membership in active_memberships:
        membership.status = MembershipStatus.REMOVED
        membership.removed_at = datetime.utcnow()
    
    # Update student count
    classroom.student_count = 0
    
    # Cancel any draft assignments
    draft_assignments = db.query(ClassroomQuizAssignment).filter(
        ClassroomQuizAssignment.classroom_id == classroom_id,
        ClassroomQuizAssignment.status == AssignmentStatus.DRAFT
    ).all()
    
    for assignment in draft_assignments:
        assignment.status = AssignmentStatus.CANCELLED
        assignment.updated_at = datetime.utcnow()
    
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to archive classroom. Please try again."
        )
    
    return {
        "message": "Classroom archived successfully",
        "classroom_id": classroom_id,
        "classroom_name": classroom.name,
        "students_removed": len(active_memberships),
        "assignments_cancelled": len(draft_assignments)
    }

# Optional: Restore archived classroom
@router.put("/{classroom_id}/restore")
async def restore_classroom(
    classroom_id: int,
    current_user: User = Depends(require_teacher),
    db: Session = Depends(get_db)
):
    """Restore an archived classroom back to active status"""
    
    # Get classroom and verify ownership
    classroom = db.query(Classroom).filter(
        Classroom.id == classroom_id,
        Classroom.teacher_id == current_user.id
    ).first()
    
    if not classroom:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Classroom not found or you don't have permission"
        )
    
    if classroom.status != ClassroomStatus.ARCHIVED:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only archived classrooms can be restored"
        )
    
    # Restore classroom
    classroom.status = ClassroomStatus.ACTIVE
    classroom.updated_at = datetime.utcnow()
    
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to restore classroom. Please try again."
        )
    
    return {
        "message": "Classroom restored successfully",
        "classroom_id": classroom_id,
        "classroom_name": classroom.name
    }

# Optional: Permanent delete (use with extreme caution)
@router.delete("/{classroom_id}/permanent")
async def permanently_delete_classroom(
    classroom_id: int,
    current_user: User = Depends(require_teacher),
    db: Session = Depends(get_db)
):
    """Permanently delete classroom and ALL related data - IRREVERSIBLE"""
    
    # Get classroom and verify ownership
    classroom = db.query(Classroom).filter(
        Classroom.id == classroom_id,
        Classroom.teacher_id == current_user.id
    ).first()
    
    if not classroom:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Classroom not found or you don't have permission"
        )
    
    # Safety check: only allow deletion if classroom is archived
    if classroom.status != ClassroomStatus.ARCHIVED:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Classroom must be archived before permanent deletion. Archive it first."
        )
    
    # Check for any student submissions (prevent data loss)
    submissions_count = db.query(ClassroomQuizSubmission).join(
        ClassroomQuizAssignment,
        ClassroomQuizSubmission.assignment_id == ClassroomQuizAssignment.id
    ).filter(
        ClassroomQuizAssignment.classroom_id == classroom_id
    ).count()
    
    if submissions_count > 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot permanently delete classroom with {submissions_count} student submissions. This would result in data loss."
        )
    
    try:
        # Due to cascade="all, delete-orphan" in the model relationships,
        # deleting the classroom should automatically delete related records
        db.delete(classroom)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to delete classroom. Please contact support."
        )
    
    return {
        "message": "Classroom permanently deleted",
        "classroom_id": classroom_id
    }

# Get archived classrooms (optional - for teachers to see what they've archived)
@router.get("/archived")
async def get_archived_classrooms(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    current_user: User = Depends(require_teacher),
    db: Session = Depends(get_db)
):
    """Get teacher's archived classrooms"""
    
    archived_classrooms = db.query(Classroom).filter(
        Classroom.teacher_id == current_user.id,
        Classroom.status == ClassroomStatus.ARCHIVED
    ).order_by(Classroom.updated_at.desc()).offset(skip).limit(limit).all()
    
    from app.schemas.classroom_schemas import ClassroomResponse
    return [
        ClassroomResponse.model_validate(classroom, from_attributes=True) 
        for classroom in archived_classrooms
    ]



@router.get("/my-classrooms", response_model=List[ClassroomResponse])
async def get_my_classrooms(
    status_filter: Optional[str] = Query(None, alias="status"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    current_user: User = Depends(require_teacher),
    db: Session = Depends(get_db)
):
    """Get teacher's classrooms"""
    service = ClassroomService(db)
    classrooms = service.get_teacher_classrooms(
        current_user.id, 
        status_filter, 
        skip, 
        limit
    )
    return [ClassroomResponse.model_validate(c , from_attributes=True) for c in classrooms]

@router.post("/join", response_model=dict)
async def join_classroom(
    join_data: ClassroomJoinRequest,
    current_user: User = Depends(require_student),
    db: Session = Depends(get_db)
):
    """Join classroom using join code (students only)"""
    service = ClassroomService(db)
    membership = service.join_classroom(current_user.id, join_data.join_code)
    
    return {
        "message": "Successfully joined classroom",
        "classroom_name": membership.classroom.name,
        "teacher_name": membership.classroom.teacher.full_name
    }

@router.get("/enrolled", response_model=List[StudentClassroomResponse])
async def get_enrolled_classrooms(
    current_user: User = Depends(require_student),
    db: Session = Depends(get_db)
):
    """Get student's enrolled classrooms"""
    service = ClassroomService(db)
    classrooms_data = service.get_student_classrooms(current_user.id)
    
    return [
       StudentClassroomResponse(
        classroom=data["classroom"],
        joined_at=data["joined_at"],
        teacher_name=data["teacher_name"],
        membership_status=data["membership_status"],  # Use from data
        completed_assignments=data["completed_assignments"],  # Use from data
        total_assignments=data["total_assignments"],  # Use from data
        average_score=data["average_score"],  # Use from data
        last_activity=data["last_activity"],  # Use from data
        pending_assignments=data["pending_assignments"]  # Add this missing field
    )
    for data in classrooms_data
    ]

@router.get("/{classroom_id}", response_model=ClassroomResponse)
async def get_classroom(
    classroom_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get classroom details"""
    service = ClassroomService(db)
    classroom = service.get_classroom_by_id(classroom_id, current_user.id)
    return ClassroomResponse.model_validate(classroom ,from_attributes=True)

@router.get("/{classroom_id}/analytics")
async def get_classroom_analytics(
    classroom_id: int,
    current_user: User = Depends(require_teacher),
    db: Session = Depends(get_db)
):
    """Get classroom analytics and performance data"""
    service = ClassroomService(db)
    return service.get_classroom_analytics(current_user.id, classroom_id)
# Quick production fix - in your router
@router.get("/{classroom_id}/students")
async def get_classroom_students(
    classroom_id: int,
    current_user: User = Depends(require_teacher),
    db: Session = Depends(get_db)
):
    try:
        service = ClassroomService(db)
        students_data = service.get_classroom_students(current_user.id, classroom_id)
        return {"students": students_data, "total": len(students_data)}
    except Exception as e:
        logger.error(f"Error fetching students: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch students")

@router.post("/{classroom_id}/assignments", response_model=AssignmentResponse)
async def assign_quiz(
    classroom_id: int,
    assignment_data: ClassroomAssignmentCreate,
    current_user: User = Depends(require_teacher),
    db: Session = Depends(get_db)
):
    """Assign quiz to classroom"""
    assignment_data.classroom_id = classroom_id
    print(f"Assignment data received: {assignment_data.dict()}")
    print(f"Geofencing enabled: {assignment_data.geofencing_enabled}")
    print(f"Coordinates: {assignment_data.allowed_latitude}, {assignment_data.allowed_longitude}")
    service = ClassroomService(db)
    assignment = service.assign_quiz_to_classroom(current_user.id, assignment_data)
    return AssignmentResponse.model_validate(assignment , from_attributes=True)

# submisson data 
@router.get("/{classroom_id}/assignments")
async def get_classroom_assignments(
    classroom_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get all assignments for classroom - with student submission data if student"""
    service = ClassroomService(db)
    
    # Check if user is a student in this classroom
    if current_user.role == UserRole.STUDENT:
        # Use the student-specific method that includes submission data
        assignments = service.get_classroom_assignments_for_student(classroom_id, current_user.id)
        return assignments
    else:
        # For teachers, use the existing method
        assignments = service.get_classroom_assignments(classroom_id, current_user.id)
        return [AssignmentResponse.model_validate(a, from_attributes=True) for a in assignments]

@router.put("/{classroom_id}", response_model=ClassroomResponse)
async def update_classroom(
    classroom_id: int,
    classroom_data: ClassroomUpdate,
    current_user: User = Depends(require_teacher),
    db: Session = Depends(get_db)
):
    """Update classroom settings"""
    service = ClassroomService(db)
    classroom = service.update_classroom(current_user.id, classroom_id, classroom_data)
    return ClassroomResponse.model_validate(classroom , from_attributes=True)

@router.delete("/{classroom_id}/students/{student_id}")
async def remove_student(
    classroom_id: int,
    student_id: int,
    current_user: User = Depends(require_teacher),
    db: Session = Depends(get_db)
):
    """Remove student from classroom"""
    service = ClassroomService(db)
    service.remove_student(current_user.id, classroom_id, student_id)
    return {"message": "Student removed successfully"}

@router.post("/{classroom_id}/leave")
async def leave_classroom(
    classroom_id: int,
    current_user: User = Depends(require_student),
    db: Session = Depends(get_db)
):
    """Student leaves classroom"""
    service = ClassroomService(db)
    service.leave_classroom(current_user.id, classroom_id)
    return {"message": "Left classroom successfully"}

# Add to classroom_routes.py

@router.get("/assignments/{assignment_id}/student-view")
async def get_assignment_for_student(
    assignment_id: int,
    current_user: User = Depends(require_student),
    db: Session = Depends(get_db)
):
    """Get assignment details for student to take quiz"""
    service = ClassroomService(db)
    assignment_data = service.get_assignment_for_student(assignment_id, current_user.id)
    return assignment_data

@router.get("/assignments/student")
async def get_student_assignments(
    current_user: User = Depends(require_student),
    db: Session = Depends(get_db)
):
    """Get all assignments for current student"""
    service = ClassroomService(db)
    assignments = service.get_student_assignments(current_user.id)
    return assignments


# Add to app/routers/classroom_routes.py (NOT assignment_submission_router.py)
# Fix 3: Add logging to the classroom route
@router.post("/{classroom_id}/assignments", response_model=AssignmentResponse)
async def assign_quiz_to_classroom(
    classroom_id: int,
    assignment_data: ClassroomAssignmentCreate,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(require_teacher),
    db: Session = Depends(get_db)
):
    """Assign quiz to classroom with enhanced notifications"""
    
    try:
        logger.info(f"Teacher {current_user.id} assigning quiz to classroom {classroom_id}")
        
        service = ClassroomService(db)
        assignment = service.assign_quiz_to_classroom(current_user.id, assignment_data)
        
        logger.info(f"Assignment created with ID: {assignment.id}")
        
        # Queue background task for notifications
        background_tasks.add_task(
            send_assignment_notifications,
            db, classroom_id, assignment.id, current_user.full_name or current_user.username
        )
        
        logger.info(f"Background task queued for assignment {assignment.id}")
        
        return AssignmentResponse.model_validate(assignment, from_attributes=True)
        
    except Exception as e:
        logger.error(f"Error in assign_quiz_to_classroom: {str(e)}")
        logger.error(f"Assignment creation traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Failed to create assignment: {str(e)}")

async def send_assignment_notifications(db: Session, classroom_id: int, assignment_id: int, teacher_name: str):
    """Enhanced assignment notifications with proper error handling"""
    
    try:
        logger.info(f"Starting assignment notifications for assignment {assignment_id}")
        
        # Fix: Use joinedload to eager load related objects
        assignment = db.query(ClassroomQuizAssignment)\
            .options(
                joinedload(ClassroomQuizAssignment.classroom),
                joinedload(ClassroomQuizAssignment.quiz)
            )\
            .filter(ClassroomQuizAssignment.id == assignment_id)\
            .first()
        
        if not assignment:
            logger.error(f"Assignment {assignment_id} not found")
            return
            
        logger.info(f"Found assignment: {assignment.title}")
        
        # Get all active students with explicit join
        students = db.query(User).join(
            ClassroomMembership, 
            User.id == ClassroomMembership.user_id
        ).filter(
            ClassroomMembership.classroom_id == classroom_id,
            ClassroomMembership.status == MembershipStatus.ACTIVE,
            ClassroomMembership.role == 'student'  # Make sure we only get students
        ).all()
        
        logger.info(f"Found {len(students)} students to notify")
        
        if not students:
            logger.warning(f"No students found in classroom {classroom_id}")
            return
        
        email_service = EmailService()
        email_success_count = 0
        notification_success_count = 0
        
        for student in students:
            try:
                logger.info(f"Processing notifications for student: {student.email}")
                
                # Create in-app notification first
                notification = Notification(
                    user_id=student.id,
                    type='quiz_assigned',
                    title=f'New Quiz: {assignment.title}',
                    message=f'{teacher_name} assigned a new quiz in {assignment.classroom.name}',
                    data={
                        'assignment_id': assignment_id,
                        'classroom_id': classroom_id,
                        'due_date': assignment.due_date.isoformat() if assignment.due_date else None
                    }
                )
                db.add(notification)
                notification_success_count += 1
                logger.info(f"Created in-app notification for {student.email}")
                
                # Send email notification
                try:
                    email_sent = email_service.send_assignment_email(
                        to_email=student.email,
                        student_name=student.full_name or student.username,
                        assignment=assignment,
                        teacher_name=teacher_name
                    )
                    
                    if email_sent:
                        email_success_count += 1
                        logger.info(f"Email sent successfully to {student.email}")
                    else:
                        logger.error(f"Failed to send email to {student.email}")
                        
                except Exception as email_error:
                    logger.error(f"Email error for {student.email}: {str(email_error)}")
                    logger.error(f"Email error traceback: {traceback.format_exc()}")
                
            except Exception as student_error:
                logger.error(f"Error processing student {student.email}: {str(student_error)}")
                logger.error(f"Student processing traceback: {traceback.format_exc()}")
        
        # Commit all notifications
        try:
            db.commit()
            logger.info(f"Database committed successfully")
        except Exception as commit_error:
            logger.error(f"Database commit failed: {str(commit_error)}")
            db.rollback()
            raise
        
        logger.info(f"Notification summary: {notification_success_count} in-app notifications, {email_success_count} emails sent out of {len(students)} students")
        
    except Exception as e:
        logger.error(f"Critical error in send_assignment_notifications: {str(e)}")
        logger.error(f"Full traceback: {traceback.format_exc()}")
        try:
            db.rollback()
        except:
            pass
        raise
    
    
    
# specific route for getting classroom details for students
@router.get("/{classroom_id}/student-view")
async def get_classroom_student_view(
    classroom_id: int,
    current_user: User = Depends(require_student),
    db: Session = Depends(get_db)
):
    """Get classroom details optimized for student view"""
    
    # Verify student access
    membership = db.query(ClassroomMembership).filter(
        ClassroomMembership.classroom_id == classroom_id,
        ClassroomMembership.user_id == current_user.id,
        ClassroomMembership.status == MembershipStatus.ACTIVE
    ).first()
    
    if not membership:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You don't have access to this classroom"
        )
    
    # Get classroom details
    classroom = db.query(Classroom).filter(
        Classroom.id == classroom_id,
        Classroom.status == ClassroomStatus.ACTIVE
    ).first()
    
    if not classroom:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Classroom not found"
        )
    
    # Get assignments count by status for this student
    all_assignments = db.query(ClassroomQuizAssignment).filter(
        ClassroomQuizAssignment.classroom_id == classroom_id,
        ClassroomQuizAssignment.status == AssignmentStatus.ACTIVE
    ).all()
    
    pending_count = 0
    completed_count = 0
    overdue_count = 0
    
    now = datetime.utcnow()
    
    for assignment in all_assignments:
        submission = db.query(ClassroomQuizSubmission).filter(
            ClassroomQuizSubmission.assignment_id == assignment.id,
            ClassroomQuizSubmission.student_id == current_user.id
        ).first()
        
        if submission:
            completed_count += 1
        elif assignment.due_date and now > assignment.due_date:
            overdue_count += 1
        else:
            pending_count += 1
    
    return {
        "classroom": {
            "id": classroom.id,
            "name": classroom.name,
            "description": classroom.description,
            "subject": classroom.subject,
            "status": classroom.status.value,
            "student_count": classroom.student_count,
            "created_at": classroom.created_at,
            "teacher_name": classroom.teacher.full_name or classroom.teacher.username,
            "teacher_id": classroom.teacher_id
        },
        "membership": {
            "joined_at": membership.joined_at,
            "role": membership.role,
            "status": membership.status.value
        },
        "assignment_summary": {
            "pending": pending_count,
            "completed": completed_count,
            "overdue": overdue_count,
            "total": len(all_assignments)
        }
    }
    
    
    
# adding for results and deletion and updation 

@router.get("/{classroom_id}/assignments/{assignment_id}/results", response_model=Dict[str, Any])
async def get_assignment_results_teacher(
    classroom_id: int,
    assignment_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get assignment results for teacher view"""
    # Verify teacher owns the classroom
    classroom = db.query(Classroom).filter(
        Classroom.id == classroom_id,
        Classroom.teacher_id == current_user.id
    ).first()
    
    if not classroom:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Classroom not found"
        )
    
    # Get assignment
    assignment = db.query(ClassroomQuizAssignment).filter(
        ClassroomQuizAssignment.id == assignment_id,
        ClassroomQuizAssignment.classroom_id == classroom_id
    ).first()
    
    if not assignment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Assignment not found"
        )
    
    # Get all submissions for this assignment
    submissions = db.query(ClassroomQuizSubmission).join(
        User, ClassroomQuizSubmission.student_id == User.id
    ).filter(
        ClassroomQuizSubmission.assignment_id == assignment_id
    ).order_by(ClassroomQuizSubmission.submitted_at.desc()).all()
    
    # Get classroom students
    students = db.query(User).join(
        ClassroomMembership, ClassroomMembership.student_id == User.id
    ).filter(
        ClassroomMembership.classroom_id == classroom_id,
        ClassroomMembership.status == MembershipStatus.ACTIVE
    ).all()
    
    # Calculate statistics
    total_students = len(students)
    submitted_count = len(set(s.student_id for s in submissions))
    pending_count = total_students - submitted_count

    # Fix: Handle empty submissions list safely
    #get_assignment_results_teacher function, fix the percentage calculation:
        # Replace the problematic lines with this safer version:
    if submissions and any(s.total_marks_scored is not None for s in submissions):
        # Calculate statistics using only submissions with valid mark data
        valid_submissions = [s for s in submissions if s.total_marks_scored is not None and s.max_possible_marks is not None]
        
        if valid_submissions:
            # Calculate average based on marks
            total_marks_earned = sum(s.total_marks_scored for s in valid_submissions)
            total_max_marks = sum(s.max_possible_marks for s in valid_submissions)
            
            if total_max_marks > 0:
                average_score = (total_marks_earned / len(valid_submissions)) / (total_max_marks / len(valid_submissions)) * 100
            else:
                average_score = 0
                
            # Calculate individual percentages for min/max
            individual_percentages = []
            for s in valid_submissions:
                if s.max_possible_marks > 0:
                    percentage = (s.total_marks_scored / s.max_possible_marks) * 100
                    individual_percentages.append(max(0, percentage))  # Cap at 0%
            
            if individual_percentages:
                highest_score = max(individual_percentages)
                lowest_score = min(individual_percentages)
            else:
                highest_score = lowest_score = 0
        else:
            average_score = highest_score = lowest_score = 0
    else:
        # Fallback to old percentage-based calculation for legacy data
        valid_scores = [s.score_percentage for s in submissions if s.score_percentage is not None]
        if valid_scores:
            average_score = sum(valid_scores) / len(valid_scores)
            highest_score = max(valid_scores)
            lowest_score = min(valid_scores)
        else:
            average_score = highest_score = lowest_score = 0
        
    # Format student results
    student_results = []
    for student in students:
        student_submission = next((s for s in submissions if s.student_id == student.id), None)
        student_results.append({
            "student_id": student.id,
            "student_name": student.full_name,
            "student_email": student.email,
            "submitted": bool(student_submission),
            "submitted_at": student_submission.submitted_at if student_submission else None,
            "score_percentage": student_submission.score_percentage if student_submission else None,
            "is_late": student_submission.is_late if student_submission else False,
            "attempt_number": student_submission.attempt_number if student_submission else 0,
            "time_taken_minutes": student_submission.time_taken_minutes if student_submission else None
        })
    
    return {
        "assignment": {
            "id": assignment.id,
            "title": assignment.title,
            "description": assignment.description,
            "due_date": assignment.due_date,
            "created_at": assignment.created_at,
            "status": assignment.status.value
        },
        "quiz": {
            "id": assignment.quiz.id,
            "title": assignment.quiz.title,
            "total_questions": assignment.quiz.total_questions,
            "total_points": assignment.quiz.total_points
        },
        "statistics": {
            "total_students": total_students,
            "submitted_count": submitted_count,
            "pending_count": pending_count,
            "completion_rate": (submitted_count / total_students * 100) if total_students > 0 else 0,
            "average_score": round(average_score, 2),
            "highest_score": round(highest_score, 2),
            "lowest_score": round(lowest_score, 2)
        },
        "student_results": student_results
    }

@router.delete("/{classroom_id}/assignments/{assignment_id}", response_model=Dict[str, Any])
async def delete_assignment(
    classroom_id: int,
    assignment_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Delete an assignment (soft delete)"""
    # Verify teacher owns the classroom
    classroom = db.query(Classroom).filter(
        Classroom.id == classroom_id,
        Classroom.teacher_id == current_user.id
    ).first()
    
    if not classroom:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Classroom not found"
        )
    
    # Get assignment
    assignment = db.query(ClassroomQuizAssignment).filter(
        ClassroomQuizAssignment.id == assignment_id,
        ClassroomQuizAssignment.classroom_id == classroom_id
    ).first()
    
    if not assignment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Assignment not found"
        )
    
    # Check if assignment has submissions
    submissions_count = db.query(ClassroomQuizSubmission).filter(
        ClassroomQuizSubmission.assignment_id == assignment_id
    ).count()
    
    if submissions_count > 0:
        # Soft delete - change status to INACTIVE (or whatever inactive status exists in your enum)
        # You can also use assignment.status = AssignmentStatus.INACTIVE if that exists
        # For now, we'll just do a soft delete by hiding it (you can modify this based on your enum)
        try:
            # Try common enum values for inactive/deleted status
            if hasattr(AssignmentStatus, 'INACTIVE'):
                assignment.status = AssignmentStatus.INACTIVE
            elif hasattr(AssignmentStatus, 'ARCHIVED'):
                assignment.status = AssignmentStatus.ARCHIVED
            elif hasattr(AssignmentStatus, 'DISABLED'):
                assignment.status = AssignmentStatus.DISABLED
            else:
                # If no suitable status exists, just hard delete
                db.delete(assignment)
                db.commit()
                return {
                    "message": "Assignment deleted successfully",
                    "assignment_id": assignment_id,
                    "action": "deleted"
                }
            
            assignment.updated_at = datetime.utcnow()
            db.commit()
            
            return {
                "message": "Assignment archived successfully (has submissions)",
                "assignment_id": assignment_id,
                "action": "archived"
            }
        except AttributeError:
            # If status update fails, do hard delete
            db.delete(assignment)
            db.commit()
            return {
                "message": "Assignment deleted successfully",
                "assignment_id": assignment_id,
                "action": "deleted"
            }
    else:
        # Hard delete if no submissions
        db.delete(assignment)
        db.commit()
        
        return {
            "message": "Assignment deleted successfully",
            "assignment_id": assignment_id,
            "action": "deleted"
        }

@router.put("/{classroom_id}/assignments/{assignment_id}", response_model=Dict[str, Any])
async def update_assignment(
    classroom_id: int,
    assignment_id: int,
    update_data: Dict[str, Any],
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update assignment details"""
    # Verify teacher owns the classroom
    classroom = db.query(Classroom).filter(
        Classroom.id == classroom_id,
        Classroom.teacher_id == current_user.id
    ).first()
    
    if not classroom:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Classroom not found"
        )
    
    # Get assignment
    assignment = db.query(ClassroomQuizAssignment).filter(
        ClassroomQuizAssignment.id == assignment_id,
        ClassroomQuizAssignment.classroom_id == classroom_id
    ).first()
    
    if not assignment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Assignment not found"
        )
    
    # Update allowed fields
    allowed_fields = [
        'title', 'description', 'instructions', 'due_date', 
        'time_limit_minutes', 'max_attempts', 'allow_late_submission',
        'show_results_immediately', 'shuffle_questions'
    ]
    
    for field, value in update_data.items():
        if field in allowed_fields and hasattr(assignment, field):
            setattr(assignment, field, value)
    
    assignment.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(assignment)
    
    return {
        "message": "Assignment updated successfully",
        "assignment": {
            "id": assignment.id,
            "title": assignment.title,
            "description": assignment.description,
            "due_date": assignment.due_date,
            "updated_at": assignment.updated_at
        }
    }
    
    
    
    
    
# This is to download  the result in the excel file and  other option 
@router.get("/{classroom_id}/assignments/{assignment_id}/export-excel")
async def export_assignment_results_excel(
    classroom_id: int,
    assignment_id: int,
    export_filter: str = "all",  # "all", "submitted", "pending"
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export assignment results to Excel file"""
    # Verify teacher owns the classroom
    classroom = db.query(Classroom).filter(
        Classroom.id == classroom_id,
        Classroom.teacher_id == current_user.id
    ).first()
    
    if not classroom:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Classroom not found"
        )
    
    # Get assignment
    assignment = db.query(ClassroomQuizAssignment).filter(
        ClassroomQuizAssignment.id == assignment_id,
        ClassroomQuizAssignment.classroom_id == classroom_id
    ).first()
    
    if not assignment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Assignment not found"
        )
    
    # Get all submissions for this assignment
    submissions = db.query(ClassroomQuizSubmission).join(
        User, ClassroomQuizSubmission.student_id == User.id
    ).filter(
        ClassroomQuizSubmission.assignment_id == assignment_id
    ).order_by(ClassroomQuizSubmission.submitted_at.desc()).all()
    
    # Get classroom students
    students = db.query(User).join(
        ClassroomMembership, ClassroomMembership.student_id == User.id
    ).filter(
        ClassroomMembership.classroom_id == classroom_id,
        ClassroomMembership.status == MembershipStatus.ACTIVE
    ).order_by(User.full_name).all()
    
    # Create submission lookup
    submission_dict = {s.student_id: s for s in submissions}
    
    # Filter students based on export_filter
    filtered_students = []
    for student in students:
        has_submitted = student.id in submission_dict
        
        if export_filter == "submitted" and not has_submitted:
            continue
        elif export_filter == "pending" and has_submitted:
            continue
        
        filtered_students.append(student)
    
    # Calculate statistics
    total_students = len(students)
    submitted_count = len(submissions)
    pending_count = total_students - submitted_count
    
        # In export_assignment_results_excel function, replace the statistics calculation:
    if submissions and any(s.total_marks_scored is not None for s in submissions):
        valid_submissions = [s for s in submissions if s.total_marks_scored is not None and s.max_possible_marks is not None]
        
        if valid_submissions:
            total_marks_earned = sum(s.total_marks_scored for s in valid_submissions)
            total_max_marks = sum(s.max_possible_marks for s in valid_submissions)
            average_score = (total_marks_earned / len(valid_submissions)) / (total_max_marks / len(valid_submissions)) * 100 if total_max_marks > 0 else 0
            
            individual_percentages = [(s.total_marks_scored / s.max_possible_marks) * 100 for s in valid_submissions if s.max_possible_marks > 0]
            highest_score = max(individual_percentages) if individual_percentages else 0
            lowest_score = min(individual_percentages) if individual_percentages else 0
        else:
            average_score = highest_score = lowest_score = 0
    else:
        # Fallback for legacy data
        valid_scores = [s.score_percentage for s in submissions if s.score_percentage is not None]
        if valid_scores:
            average_score = sum(valid_scores) / len(valid_scores)
            highest_score = max(valid_scores)
            lowest_score = min(valid_scores)
        else:
            average_score = highest_score = lowest_score = 0
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Assignment Results"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(bold=True, size=16)
    subtitle_font = Font(bold=True, size=12)
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title and metadata
    ws.merge_cells('A1:H1')
    ws['A1'] = f"{assignment.title} - Results Report"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    
    ws.merge_cells('A2:H2')
    ws['A2'] = f"Class: {classroom.name} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].alignment = center_align
    
    # Summary statistics
    current_row = 4
    ws[f'A{current_row}'] = "Summary Statistics"
    ws[f'A{current_row}'].font = subtitle_font
    current_row += 1
    
    stats_data = [
        ("Total Students", total_students),
        ("Submitted", submitted_count),
        ("Pending", pending_count),
        ("Completion Rate", f"{(submitted_count/total_students*100):.1f}%" if total_students > 0 else "0%"),
        ("Average Score", f"{average_score:.1f}%" if average_score > 0 else "N/A"),
        ("Highest Score", f"{highest_score:.1f}%" if highest_score > 0 else "N/A"),
        ("Lowest Score", f"{lowest_score:.1f}%" if lowest_score > 0 else "N/A")
    ]
    
    for label, value in stats_data:
        ws[f'A{current_row}'] = label
        ws[f'B{current_row}'] = value
        ws[f'A{current_row}'].font = Font(bold=True)
        current_row += 1
    
    # Student results section
    current_row += 2
    ws[f'A{current_row}'] = "Student Results"
    ws[f'A{current_row}'].font = subtitle_font
    current_row += 1
    
    # Headers
    headers = ["Name", "Email", "Username", "Score (Marks)", "Score (%)", "Status", "Submitted At", "Late", "Time Taken (min)", "Attempts"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    current_row += 1
    
    # Student data
    for student in filtered_students:
        submission = submission_dict.get(student.id)
        
        # Format score display
        if submission:
            if submission.total_marks_scored is not None and submission.max_possible_marks is not None:
                # New format: show actual marks
                marks_display = f"{submission.total_marks_scored}/{submission.max_possible_marks}"
                percentage_display = f"{(submission.total_marks_scored / submission.max_possible_marks * 100):.1f}%" if submission.max_possible_marks > 0 else "0%"
            else:
                # Fallback for legacy data
                marks_display = "N/A"
                percentage_display = f"{submission.score_percentage:.1f}%" if submission.score_percentage is not None else "Not Graded"
        else:
            marks_display = "Not Submitted"
            percentage_display = "Not Submitted"
        
        row_data = [
            student.full_name or "N/A",
            student.email or "N/A", 
            student.username or "N/A",
            marks_display,  # Actual marks column
            percentage_display,  # Percentage column
            "Submitted" if submission else "Pending",
            submission.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if submission else "N/A",
            "Yes" if submission and submission.is_late else "No",
            str(submission.time_taken_minutes) if submission and submission.time_taken_minutes else "N/A",
            str(submission.attempt_number) if submission else "0"
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = value
            cell.border = border
            cell.alignment = center_align
            
            # Color coding for marks column (col 4) and percentage column (col 5)
            if col == 4 and submission and submission.total_marks_scored is not None:
                # Color based on whether marks are positive/negative
                if submission.total_marks_scored < 0:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red for negative
                elif submission.total_marks_scored >= submission.max_possible_marks * 0.8:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green for 80%+
                elif submission.total_marks_scored >= submission.max_possible_marks * 0.6:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow for 60-79%
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red for <60%
            
            elif col == 5 and submission and submission.total_marks_scored is not None:
                # Color coding for percentage column
                percentage = (submission.total_marks_scored / submission.max_possible_marks * 100) if submission.max_possible_marks > 0 else 0
                if percentage >= 80:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
                elif percentage >= 60:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
        
        current_row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to memory
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create filename (sanitize for cross-platform compatibility)
    safe_assignment_title = re.sub(r'[^\w\s-]', '', assignment.title)
    safe_assignment_title = re.sub(r'[-\s]+', '-', safe_assignment_title)
    date_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"Assignment_Results_{safe_assignment_title}_{date_str}.xlsx"
    
    # Return streaming response
    def iter_file():
        excel_buffer.seek(0)
        yield from excel_buffer
    
    return StreamingResponse(
        iter_file(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
    
    
    
# for downloading the excel file for each class their children score 

@router.get("/{classroom_id}/export-classroom-excel")
async def export_classroom_performance_excel(
    classroom_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export all students' performance across all assignments in Excel format"""
    # Verify teacher owns the classroom
    classroom = db.query(Classroom).filter(
        Classroom.id == classroom_id,
        Classroom.teacher_id == current_user.id
    ).first()
    
    if not classroom:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Classroom not found"
        )
    
    # Get all students in classroom
    students = db.query(User).join(
        ClassroomMembership, ClassroomMembership.student_id == User.id
    ).filter(
        ClassroomMembership.classroom_id == classroom_id,
        ClassroomMembership.status == MembershipStatus.ACTIVE
    ).order_by(User.full_name).all()
    
    # Get all assignments for this classroom
    assignments = db.query(ClassroomQuizAssignment).filter(
        ClassroomQuizAssignment.classroom_id == classroom_id
    ).order_by(ClassroomQuizAssignment.created_at).all()
    
    # Get all submissions for this classroom
    submissions = db.query(ClassroomQuizSubmission).join(
        ClassroomQuizAssignment,
        ClassroomQuizSubmission.assignment_id == ClassroomQuizAssignment.id
    ).filter(
        ClassroomQuizAssignment.classroom_id == classroom_id
    ).all()
    
    # Create submission lookup: {student_id: {assignment_id: submission}}
    submission_lookup = {}
    for submission in submissions:
        if submission.student_id not in submission_lookup:
            submission_lookup[submission.student_id] = {}
        submission_lookup[submission.student_id][submission.assignment_id] = submission
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Classroom Performance"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(bold=True, size=16)
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells(f'A1:{get_column_letter(len(assignments) + 6)}1')
    ws['A1'] = f"{classroom.name} - Complete Performance Report"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    
    # Classroom info
    ws.merge_cells(f'A2:{get_column_letter(len(assignments) + 6)}2')
    ws['A2'] = f"Teacher: {classroom.teacher.full_name} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].alignment = center_align
    
    # Headers row
    current_row = 4
    headers = ["Student Name", "Email", "Username"]
    
    # Add assignment titles as headers
    for assignment in assignments:
        headers.append(f"{assignment.title}")
    
    headers.extend(["Overall Average", "Total Assignments", "Completed"])
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    current_row += 1
    
    # Student data
    for student in students:
        student_scores = []
        completed_count = 0
        
        row_data = [
            student.full_name or "N/A",
            student.email or "N/A",
            student.username or "N/A"
        ]
        
        # Add scores for each assignment
        for assignment in assignments:
            submission = submission_lookup.get(student.id, {}).get(assignment.id)
            
            if submission:
                if submission.total_marks_scored is not None and submission.max_possible_marks is not None:
                    # Show marks format: "8/10 (80%)"
                    marks_display = f"{submission.total_marks_scored}/{submission.max_possible_marks}"
                    percentage = (submission.total_marks_scored / submission.max_possible_marks * 100) if submission.max_possible_marks > 0 else 0
                    score_display = f"{marks_display} ({percentage:.1f}%)"
                    student_scores.append(percentage)
                else:
                    # Fallback to percentage only
                    score_display = f"{submission.score_percentage:.1f}%" if submission.score_percentage else "Graded"
                    student_scores.append(submission.score_percentage or 0)
                completed_count += 1
            else:
                score_display = "Not Submitted"
                student_scores.append(0)  # Count as 0 for average calculation
            
            row_data.append(score_display)
        
        # Calculate overall average
        overall_average = sum(student_scores) / len(student_scores) if student_scores else 0
        
        # Add summary columns
        row_data.extend([
            f"{overall_average:.1f}%",
            len(assignments),
            completed_count
        ])
        
        # Write student row
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = value
            cell.border = border
            cell.alignment = center_align
            
            # Color coding for assignment scores (columns 4 to 4+len(assignments)-1)
            if 4 <= col <= 3 + len(assignments):
                if "Not Submitted" in str(value):
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
                elif any(char.isdigit() for char in str(value)):
                    # Extract percentage for color coding
                    if "(" in str(value) and "%" in str(value):
                        try:
                            percentage_str = str(value).split("(")[1].split("%")[0]
                            percentage = float(percentage_str)
                            if percentage >= 80:
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
                            elif percentage >= 60:
                                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
                            else:
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
                        except:
                            pass
        
        current_row += 1
    
    # Add summary statistics
    current_row += 2
    ws[f'A{current_row}'] = "Classroom Statistics"
    ws[f'A{current_row}'].font = Font(bold=True, size=14)
    current_row += 1
    
    # Calculate class averages for each assignment
    for i, assignment in enumerate(assignments):
        assignment_scores = []
        for student in students:
            submission = submission_lookup.get(student.id, {}).get(assignment.id)
            if submission and submission.score_percentage:
                assignment_scores.append(submission.score_percentage)
        
        avg_score = sum(assignment_scores) / len(assignment_scores) if assignment_scores else 0
        completion_rate = len(assignment_scores) / len(students) * 100 if students else 0
        
        ws[f'A{current_row}'] = f"{assignment.title} Average:"
        ws[f'B{current_row}'] = f"{avg_score:.1f}% ({completion_rate:.1f}% completion)"
        current_row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to memory
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create filename
    safe_classroom_name = re.sub(r'[^\w\s-]', '', classroom.name)
    safe_classroom_name = re.sub(r'[-\s]+', '-', safe_classroom_name)
    date_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"Classroom_Performance_{safe_classroom_name}_{date_str}.xlsx"
    
    # Return streaming response
    def iter_file():
        excel_buffer.seek(0)
        yield from excel_buffer
    
    return StreamingResponse(
        iter_file(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )